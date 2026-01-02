"""
Reader v3 - Leitura de arquivos Excel/CSV com limpeza automática
Melhorias:
- Integração com módulo cleaner para remoção de linhas vazias, logos e rodapés
- Melhor tratamento de células mescladas
- Suporte a diferentes encodings de CSV
"""
from io import BytesIO
from typing import Dict

import pandas as pd
from openpyxl import load_workbook

from app.cleaner import clean_spreadsheet


def read_excel(file) -> Dict[str, pd.DataFrame]:
    """
    Lê arquivo Excel ou CSV e retorna dict com DataFrames por sheet.
    
    AGORA COM LIMPEZA AUTOMÁTICA:
    - Remove linhas vazias do topo
    - Remove logos/imagens
    - Remove rodapés desnecessários
    
    Para Excel: preserva informação de células mescladas.
    Para CSV: retorna com chave "CSV".
    """
    filename = (getattr(file, "filename", "") or "").lower()
    
    # Usa o cleaner para obter planilhas limpas
    cleaned_sheets = clean_spreadsheet(file, filename)
    
    # Se for Excel, ainda processa células mescladas
    if not filename.endswith(".csv"):
        # Recarrega para processar merged cells
        file.file.seek(0)
        content = file.file.read()
        merged_sheets = _read_excel_with_merged_cells(content)
        
        # Combina: usa os dados limpos mas preserva merged info
        for sheet_name in cleaned_sheets.keys():
            if sheet_name in merged_sheets:
                # Aplica limpeza nos dados com merged cells
                df = merged_sheets[sheet_name]
                # Remove linhas vazias do topo
                first_data_row = 0
                for idx in range(len(df)):
                    if df.iloc[idx].notna().any():
                        first_data_row = idx
                        break
                df = df.iloc[first_data_row:].reset_index(drop=True)
                
                # Remove rodapés
                cutoff = len(df)
                check_range = min(15, len(df))
                for i in range(len(df) - 1, max(0, len(df) - check_range - 1), -1):
                    row = df.iloc[i]
                    fill_rate = row.notna().sum() / len(row) if len(row) > 0 else 0
                    if fill_rate < 0.2:
                        cutoff = i
                    else:
                        break
                df = df.iloc[:cutoff].reset_index(drop=True)
                
                cleaned_sheets[sheet_name] = df
    
    return cleaned_sheets


def _read_csv(content: bytes) -> Dict[str, pd.DataFrame]:
    """Lê arquivo CSV com detecção de encoding."""
    # Tenta diferentes encodings
    encodings = ['utf-8', 'latin-1', 'cp1252', 'iso-8859-1']
    
    for encoding in encodings:
        try:
            df = pd.read_csv(
                BytesIO(content), 
                header=None, 
                encoding=encoding,
                on_bad_lines='skip',
                low_memory=False
            )
            return {"CSV": df}
        except Exception:
            continue
    
    # Fallback: força utf-8 ignorando erros
    df = pd.read_csv(
        BytesIO(content), 
        header=None, 
        encoding='utf-8', 
        errors='ignore',
        on_bad_lines='skip',
        low_memory=False
    )
    return {"CSV": df}


def _read_excel_with_merged_cells(content: bytes) -> Dict[str, pd.DataFrame]:
    """Lê arquivo Excel preservando merged cells."""
    try:
        wb = load_workbook(BytesIO(content), data_only=True)
        sheets = {}

        for ws in wb.worksheets:
            max_row = ws.max_row or 0
            max_col = ws.max_column or 0

            # Cria grid vazio
            grid = [
                [ws.cell(row=r, column=c).value for c in range(1, max_col + 1)]
                for r in range(1, max_row + 1)
            ]

            # Preenche células mescladas
            for merged in ws.merged_cells.ranges:
                min_col, min_row, max_col_m, max_row_m = merged.bounds
                
                # Pega valor do canto superior esquerdo
                if (min_row - 1) < len(grid) and (min_col - 1) < len(grid[min_row - 1]):
                    top_left = grid[min_row - 1][min_col - 1]
                else:
                    continue
                
                if top_left is None:
                    continue

                # Propaga para todas as células do range
                for rr in range(min_row, max_row_m + 1):
                    for cc in range(min_col, max_col_m + 1):
                        if rr == min_row and cc == min_col:
                            continue  # Já tem o valor
                        if rr - 1 >= len(grid):
                            continue
                        row = grid[rr - 1]
                        if cc - 1 >= len(row):
                            continue
                        if row[cc - 1] is None:
                            row[cc - 1] = top_left

            sheets[ws.title] = pd.DataFrame(grid)

        return sheets
        
    except Exception as e:
        # Fallback: usa pandas direto (pode perder merge info)
        try:
            excel = pd.read_excel(
                BytesIO(content),
                sheet_name=None,
                header=None,
                engine="openpyxl",
            )

            sheets = {}
            for name, df in excel.items():
                sheets[name] = df

            return sheets
        except Exception as e2:
            raise ValueError(f"Não foi possível ler o arquivo Excel: {e2}")
